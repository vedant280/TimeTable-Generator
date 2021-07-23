#import statements
#Project is made by assuming 6 subjects having labs and theory lecture 
#In working days from Monday to Friday in Self assumed working hours  
import openpyxl
import random 
from tkinter import*
from tkinter import Frame,ttk
from tkinter import messagebox as mbox

workbook = openpyxl.Workbook()
sheet = workbook.active

workbook2 = openpyxl.Workbook()
sheet2 = workbook2.active

win=Tk()
win.title('                                                                                                                                                                INDIAN INSTITUTE OF INFORMATION TECHNOLOGY , NAGPUR ')
win.configure(bg='#BC8F8F')
win.geometry('1500x650')

#labels & entrys
# for subjects 
subject1_label=Label(win,text='Subject Name',bg='#BC8F8F',fg="white",font=("Times New roman",13,'bold'))
subject1_label.grid(row=0 ,column=0,sticky=W,padx=30,pady=10)
subject1_name=StringVar()
subject1_entrybox=Entry(win,width=20,textvariable=subject1_name,bd=4,relief=RIDGE)
subject1_entrybox.grid(row=0,column=1)


subject2_label=Label(win,text='Subject Name',bg='#BC8F8F',fg="white",font=("Times New roman",13,'bold'))
subject2_label.grid(row=1 ,column=0,sticky=W,padx=30,pady=12)
subject2_name=StringVar()
subject2_entrybox=Entry(win,width=20,textvariable=subject2_name,bd=4,relief=RIDGE)
subject2_entrybox.grid(row=1,column=1)



subject3_label=Label(win,text='Subject Name',bg='#BC8F8F',fg="white",font=("Times New roman",13,'bold'))
subject3_label.grid(row=2 ,column=0,sticky=W,padx=30,pady=10)
subject3_name=StringVar()
subject3_entrybox=Entry(win,width=20,textvariable=subject3_name,bd=4,relief=RIDGE)
subject3_entrybox.grid(row=2,column=1)


subject4_label=Label(win,text='Subject Name',bg='#BC8F8F',fg="white",font=("Times New roman",13,'bold'))
subject4_label.grid(row=3 ,column=0,sticky=W,padx=30,pady=10)
subject4_name=StringVar()
subject4_entrybox=Entry(win,width=20,textvariable=subject4_name,bd=4,relief=RIDGE)
subject4_entrybox.grid(row=3,column=1)


subject5_label=Label(win,text='Subject Name',bg='#BC8F8F',fg="white",font=("Times New roman",13,'bold'))
subject5_label.grid(row=4 ,column=0,sticky=W,padx=30,pady=10)
subject5_name=StringVar()
subject5_entrybox=Entry(win,width=20,textvariable=subject5_name,bd=4,relief=RIDGE)
subject5_entrybox.grid(row=4,column=1)


subject6_label=Label(win,text='Subject Name',bg='#BC8F8F',fg="white",font=("Times New roman",13,'bold'))
subject6_label.grid(row=5 ,column=0,sticky=W,padx=30,pady=10)
subject6_name=StringVar()
subject6_entrybox=Entry(win,width=20,textvariable=subject6_name,bd=4,relief=RIDGE)
subject6_entrybox.grid(row=5,column=1)

#Teacher labels 

teacher1_label=Label(win,text='Proffesors Name',bg='#BC8F8F',fg="white",font=("Times New roman",13,'bold'))
teacher1_label.grid(row=0 ,column=2,sticky=W,padx=30,pady=10)
teacher1_name=StringVar()
teacher1_entrybox=Entry(win,width=30,textvariable=teacher1_name,bd=4,relief=RIDGE)
teacher1_entrybox.grid(row=0,column=3,sticky=W)

teacher2_label=Label(win,text='Proffesors Name',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
teacher2_label.grid(row=1 ,column=2,sticky=W,padx=30,pady=10)
teacher2_name=StringVar()
teacher2_entrybox=Entry(win,width=30,textvariable=teacher2_name,bd=4,relief=RIDGE)
teacher2_entrybox.grid(row=1,column=3,sticky=W)


teacher3_label=Label(win,text='Proffesors Name',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
teacher3_label.grid(row=2 ,column=2,sticky=W,padx=30,pady=10)
teacher3_name=StringVar()
teacher3_entrybox=Entry(win,width=30,textvariable=teacher3_name,bd=4,relief=RIDGE)
teacher3_entrybox.grid(row=2,column=3,sticky=W)


teacher4_label=Label(win,text='Proffesors Name',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
teacher4_label.grid(row=3 ,column=2,sticky=W,padx=30,pady=10)
teacher4_name=StringVar()
teacher4_entrybox=Entry(win,width=30,textvariable=teacher4_name,bd=4,relief=RIDGE)
teacher4_entrybox.grid(row=3,column=3,sticky=W)


teacher5_label=Label(win,text='Proffesors Name',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
teacher5_label.grid(row=4 ,column=2,sticky=W,padx=30,pady=10)
teacher5_name=StringVar()
teacher5_entrybox=Entry(win,width=30,textvariable=teacher5_name,bd=4,relief=RIDGE)
teacher5_entrybox.grid(row=4,column=3,sticky=W)



teacher6_label=Label(win,text='Proffesors Name',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
teacher6_label.grid(row=5 ,column=2,sticky=W,padx=30,pady=10)
teacher6_name=StringVar()
teacher6_entrybox=Entry(win,width=30,textvariable=teacher6_name,bd=4,relief=RIDGE)
teacher6_entrybox.grid(row=5,column=3,sticky=W)




theory1_label=Label(win,text='Theory Lectures Per Week',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
theory1_label.grid(row=0 ,column=3,padx=220,sticky=W,pady=10)
theory1_name=IntVar()
theory1_entrybox=Entry(win,width=7,textvariable=theory1_name,bd=4,relief=RIDGE)
theory1_entrybox.grid(row=0,column=3,padx=480,pady=10,sticky=W)

theory2_label=Label(win,text='Theory Lectures Per Week',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
theory2_label.grid(row=1 ,column=3,padx=220,sticky=W)
theory2_name=IntVar()
theory2_entrybox=Entry(win,width=7,textvariable=theory2_name,bd=4,relief=RIDGE)
theory2_entrybox.grid(row=1,column=3,padx=480,pady=10,sticky=W)

theory3_label=Label(win,text='Theory Lectures Per Week',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
theory3_label.grid(row=2 ,column=3,padx=220,sticky=W)
theory3_name=IntVar()
theory3_entrybox=Entry(win,width=7,textvariable=theory3_name,bd=4,relief=RIDGE)
theory3_entrybox.grid(row=2,column=3,padx=480,pady=10,sticky=W)

theory4_label=Label(win,text='Theory Lectures Per Week',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
theory4_label.grid(row=3 ,column=3,padx=220,sticky=W)
theory4_name=IntVar()
theory4_entrybox=Entry(win,width=7,textvariable=theory4_name,bd=4,relief=RIDGE)
theory4_entrybox.grid(row=3,column=3,padx=480,pady=10,sticky=W)

theory5_label=Label(win,text='Theory Lectures Per Week',font=("Times New roman",13,'bold'),bg='#BC8F8F',fg="white")
theory5_label.grid(row=4 ,column=3,padx=220,sticky=W)
theory5_name=IntVar()
theory5_entrybox=Entry(win,width=7,textvariable=theory5_name,bd=4,relief=RIDGE)
theory5_entrybox.grid(row=4,column=3,padx=480,pady=10,sticky=W)

theory6_label=Label(win,text='Theory Lectures Per Week',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
theory6_label.grid(row=5 ,column=3,padx=220,sticky=W)
theory6_name=IntVar()
theory6_entrybox=Entry(win,width=7,textvariable=theory6_name,bd=4,relief=RIDGE)
theory6_entrybox.grid(row=5,column=3,padx=480,pady=10,sticky=W)



lab1_label=Label(win,text='Lab Sessions Per Week',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
lab1_label.grid(row=0 ,column=3,sticky=W,padx=590,pady=20)
lab1_name=IntVar()
lab1_combobox=ttk.Combobox(win,width=7,textvariable=lab1_name,state='readonly')
lab1_combobox['values']=(0,1,2)
lab1_combobox.grid(row=0,column=3,padx=800,sticky=W)
lab1_combobox.current(0)

lab2_label=Label(win,text='Lab Sessions Per Week',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
lab2_label.grid(row=1 ,column=3,sticky=W,padx=590,pady=20)
lab2_name=IntVar()
lab2_combobox=ttk.Combobox(win,width=7,textvariable=lab2_name,state='readonly')
lab2_combobox['values']=(0,1,2)
lab2_combobox.grid(row=1,column=3,padx=800,sticky=W)
lab2_combobox.current(0)

lab3_label=Label(win,text='Lab Sessions Per Week',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
lab3_label.grid(row=2 ,column=3,sticky=W,padx=590,pady=20)
lab3_name=IntVar()
lab3_combobox=ttk.Combobox(win,width=7,textvariable=lab3_name,state='readonly')
lab3_combobox['values']=(0,1,2)
lab3_combobox.grid(row=2,column=3,padx=800,sticky=W)
lab3_combobox.current(0)


lab4_label=Label(win,text='Lab Sessions Per Week',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
lab4_label.grid(row=3 ,column=3,sticky=W,padx=590,pady=20)
lab4_name=IntVar()
lab4_combobox=ttk.Combobox(win,width=7,textvariable=lab4_name,state='readonly')
lab4_combobox['values']=(0,1,2)
lab4_combobox.grid(row=3,column=3,padx=800,sticky=W)
lab4_combobox.current(0)

lab5_label=Label(win,text='Lab Sessions Per Week',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
lab5_label.grid(row=4 ,column=3,sticky=W,padx=590,pady=20)
lab5_name=IntVar()
lab5_combobox=ttk.Combobox(win,width=7,textvariable=lab5_name,state='readonly')
lab5_combobox['values']=(0,1,2)
lab5_combobox.grid(row=4,column=3,padx=800,sticky=W)
lab5_combobox.current(0)

lab6_label=Label(win,text='Lab Sessions Per Week',fg="white",font=("Times New roman",13,'bold'),bg='#BC8F8F')
lab6_label.grid(row=5 ,column=3,sticky=W,padx=590,pady=20)
lab6_name=IntVar()
lab6_combobox=ttk.Combobox(win,width=7,textvariable=lab6_name,state='readonly')
lab6_combobox['values']=(0,1,2)
lab6_combobox.grid(row=5,column=3,padx=800,sticky=W)
lab6_combobox.current(0)

#action performed after hitting submit button
def SUBMIT():
    sub=[subject1_name.get(),subject2_name.get(),subject3_name.get(),subject4_name.get(),subject5_name.get(),subject6_name.get()]

    teacher=[teacher1_name.get(),teacher2_name.get(),teacher3_name.get(),teacher4_name.get(),teacher5_name.get(),teacher6_name.get()]

    number_class=[theory1_name.get(),theory2_name.get(),theory3_name.get(),theory4_name.get(),theory5_name.get(),theory6_name.get()]

    lab=[lab1_name.get(),lab2_name.get(),lab3_name.get(),lab4_name.get(),lab5_name.get(),lab6_name.get()]

    total_subject = []
    lab_sub = []

    for i in range(0,6):
        for j in range(0,number_class[i]):
             total_subject.append(sub[i])
    
    for i in range(0,6):
        for j in range(0,lab[i]):
            lab_sub.append(sub[i]+'_lab')
    
    random.shuffle(total_subject)
    random.shuffle(lab_sub)

    total = sum(number_class)
    tot_lab = sum(lab)

    pre_time = {
    1 : ['0','0','0','0','B','1','1','1'],
    2 : ['0','0','0','0','R','1','1','1'],
    3 : ['0','0','0','0','E','1','1','1'],
    4 : ['0','0','0','0','A','1','1','1'],
    5 : ['0','0','0','0','K','1','1','1'],
    
    }
####self assumed time for working hours 
    timetable = {
        0 : ['8:0-9:0','9:0-10:0','10:0-11:0','11:0-12:0','12:0-13:0','13:0-14:0','14:0-15:0'],
        1 : ['0','0','0','0','B','1','1','1'],
        2 : ['0','0','0','0','R','1','1','1'],
        3 : ['0','0','0','0','E','1','1','1'],
        4 : ['0','0','0','0','A','1','1','1'],
        5 : ['0','0','0','0','K','1','1','1'],
    }
    

    i = 0
    count = 1
    while(i != tot_lab):
        if count == 6: count=1
        if lab_sub[i] not in timetable[count]:
            if '1' in pre_time[count]:
                x = ''.join(pre_time[count]).index('1')
                pre_time[count][x] = 'x'
                timetable[count][x] = lab_sub[i]
                i += 1
        count += 1
        
    i = 0
    count = 1
    while(i != total):
        if count == 6: count=1
        if total_subject[i] not in timetable[count]:
            if '0' in pre_time[count]:
                x = ''.join(pre_time[count]).rindex('0')
                pre_time[count][x] = 'x'
                timetable[count][x] = total_subject[i]
                i += 1
        count += 1
        

    row = 1
    lis = ['Monday','Tuesday','Wednesday','Thrusday','Friday','Day']

    for key,values in timetable.items():
        sheet.cell(row=row, column=1, value=lis[key-1])
        column = 2
        for element in values:
            if(element != '0' and element !='1'): sheet.cell(row=row, column=column, value=element)
            column += 1
        row += 1
    
    pre_time = {
        1 : ['0','0','0','0','B','1','1','1'],
        2 : ['0','0','0','0','R','1','1','1'],
        3 : ['0','0','0','0','E','1','1','1'],
        4 : ['0','0','0','0','A','1','1','1'],
        5 : ['0','0','0','0','K','1','1','1'],
    }
    
    timetable2 = {
        0 : ['8:0-9:0','9:0-10:0','10:0-11:0','11:0-12:0','12:0-13:0','13:0-14:0','14:0-15:0'],
        1 : ['0','0','0','0','B','1','1','1'],
        2 : ['0','0','0','0','R','1','1','1'],
        3 : ['0','0','0','0','E','1','1','1'],
        4 : ['0','0','0','0','A','1','1','1'],
        5 : ['0','0','0','0','K','1','1','1'],
    }
    
    lab_sub.reverse()
        
    i = 0
    count = 1
    while(i != tot_lab):
        if count == 6: count=1
        if lab_sub[i] not in timetable2[count]:
            if '1' in pre_time[count]:
                x = ''.join(pre_time[count]).index('1')
                pre_time[count][x] = 'x'
                timetable2[count][x] = lab_sub[i]
                i += 1
        count += 1
        
    total_subject.reverse()
    i = 0
    count = 1
    while(i != total):
        if count == 6: count=1
        if total_subject[i] not in timetable2[count]:
            if '0' in pre_time[count]:
                x = ''.join(pre_time[count]).rindex('0')
                pre_time[count][x] = 'x'
                timetable2[count][x] = total_subject[i]
                i += 1
        count += 1
            

    row = 1

    for key,values in timetable2.items():
        sheet2.cell(row=row, column=1, value=lis[key-1])
        column = 2
        for element in values:
            if(element != '0' and element !='1'): sheet2.cell(row=row, column=column, value=element)
            column += 1
        row += 1


    labo=sum(lab)
    if labo==0:
        mbox.showerror('ERROR','Please fill laboratory Hours')
    if total==0:
        mbox.showerror('ERROR','Please fill Theory Hours')

##generates timetable for two shifts i
    workbook2.save(filename="timetable_A.xlsx")
    workbook.save(filename="timetable_B.xlsx")


    subject1_entrybox.delete(0,END)
    subject2_entrybox.delete(0,END)
    subject3_entrybox.delete(0,END)
    subject4_entrybox.delete(0,END)
    subject5_entrybox.delete(0,END)
    subject6_entrybox.delete(0,END)

    teacher1_entrybox.delete(0,END)
    teacher2_entrybox.delete(0,END)
    teacher3_entrybox.delete(0,END)
    teacher4_entrybox.delete(0,END)
    teacher5_entrybox.delete(0,END)
    teacher6_entrybox.delete(0,END)

    theory1_entrybox.delete(0,END)
    theory2_entrybox.delete(0,END)
    theory3_entrybox.delete(0,END)
    theory4_entrybox.delete(0,END)
    theory5_entrybox.delete(0,END)
    theory6_entrybox.delete(0,END)

    lab1_combobox.delete(0,END)
    lab2_combobox.delete(0,END)
    lab3_combobox.delete(0,END)
    lab4_combobox.delete(0,END)
    lab5_combobox.delete(0,END)
    lab6_combobox.delete(0,END)

submit_button=Button(win,bd=5,text=' SUBMIT ',width=15 ,relief=RAISED,height=2 ,bg="#20B2AA",fg="gold" ,font=("times new roman",18,"bold") ,command=SUBMIT)
submit_button.grid(row=20,column=3 ,padx=80,pady=80,sticky=W)

win.mainloop()